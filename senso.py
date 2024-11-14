import sys
from PyQt6.QtWidgets import QTableWidgetItem, QFileDialog, QComboBox, QApplication, QTableWidget, QVBoxLayout, QPushButton, QWidget, QFrame, QScrollArea, QHBoxLayout, QSpacerItem, QSizePolicy, QLabel
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt, QSettings, QStandardPaths
from openpyxl import Workbook
from PyQt6.QtGui import QGuiApplication
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import datetime
import re
from openpyxl import load_workbook
from PyQt6.QtCore import QDateTime

class Ui_Form(object):
    def setupUi(self, Form, frame, tela):
        self.tela = tela
        self.frame = frame
        self.settings = QSettings('HC', 'SGL')
        self.form = Form
        script_directory = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.AppDataLocation)
        config_file_path = f'{script_directory}/config.ini'
        self.settings = QSettings(config_file_path, QSettings.Format.IniFormat)
        self.tabela_alt = QtWidgets.QTableWidget()
        self.tabela_alt.setFixedSize(1383, 680)
        screen = QGuiApplication.primaryScreen()
        size = screen.size()
        width = size.width() - 10
        height = size.height() - 100
        self.scroll_area = QScrollArea(self.frame)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.move(0, 0)
        self.scroll_area.show()
        self.scroll_area.setStyleSheet('border:none;')
        self.scroll_area.setFixedSize(width, height)
        self.scroll_content = QWidget(self.scroll_area)
        self.scroll_area.setWidget(self.scroll_content)
        self.scroll_layout = QHBoxLayout(self.scroll_content)
        self.scroll_content.setLayout(self.scroll_layout)
        self.frames = []
        folder_path = 'senso'
        self.day = 'ONTEM'
        from datetime import datetime
        current_month = datetime.now().strftime('%B')
        self.atualizar_options_data()
        senso_folder = os.path.join(os.getcwd(), 'senso')
        current_month_year = datetime.now().strftime('%B_%Y')
        month_year_folder = os.path.join(senso_folder, current_month_year)
        self.load_tables_from_folder(month_year_folder)
        self.tela.btn_filtros.setFocus()
        current_datetime = QDateTime.currentDateTime()
        formatted_date = current_datetime.toString('yyyy')
        formatted_date2 = current_datetime.addYears((-1)).toString('yyyy')
        ccurrent_datetime = QtCore.QDateTime.currentDateTime()
        self.borda_inicio = QtWidgets.QFrame(parent=self.tela.frame_personalisa)
        self.borda_inicio.setStyleSheet('border: 2px solid black; border-radius: 10px; background-color: #FFFFFF;')
        self.borda_inicio.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.borda_inicio.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.borda_inicio.setObjectName('frame')
        self.borda_inicio.setGeometry(QtCore.QRect(8, 25, 140, 29))
        self.data_inicio = QtWidgets.QDateEdit(parent=self.tela.frame_personalisa)
        self.data_inicio.setGeometry(QtCore.QRect(10, 30, 130, 20))
        self.data_inicio.setDateTime(ccurrent_datetime)
        self.data_inicio.setCalendarPopup(True)
        self.data_inicio.setStyleSheet('border_color: #FFFFFF; border-radius: 10px; background-color: #FFFFFF;')
        self.data_inicio.setObjectName('data_inicio')
        self.inicio = QtWidgets.QLabel('Depois de ', self.tela.frame_personalisa)
        self.inicio.setStyleSheet('font-size: 13px; margin: 0; padding: 0;border: none; background-color: #FFFFFF')
        self.inicio.setGeometry(15, 20, 63, 13)
        self.borda_fim = QtWidgets.QFrame(parent=self.tela.frame_personalisa)
        self.borda_fim.setStyleSheet('border: 2px solid black; border-radius: 10px; background-color: #FFFFFF;')
        self.borda_fim.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        self.borda_fim.setFrameShadow(QtWidgets.QFrame.Shadow.Raised)
        self.borda_fim.setObjectName('frame')
        self.borda_fim.setGeometry(QtCore.QRect(8, 80, 140, 29))
        self.data_final = QtWidgets.QDateEdit(parent=self.tela.frame_personalisa)
        self.data_final.setGeometry(QtCore.QRect(10, 85, 130, 20))
        self.data_final.setStyleSheet('border_color: #FFFFFF; border-radius: 10px; background-color: #FFFFFF;')
        self.data_final.setDateTime(QtCore.QDateTime.currentDateTime())
        self.data_final.setCalendarPopup(True)
        self.data_final.setObjectName('data_inicio')
        self.fim = QtWidgets.QLabel('Antes de ', self.tela.frame_personalisa)
        self.fim.setStyleSheet('font-size: 13px; margin: 0; padding: 0;border: none; background-color: #FFFFFF')
        self.fim.setGeometry(15, 75, 57, 13)
        self.aplicar = QtWidgets.QPushButton('Aplicar', parent=self.tela.frame_personalisa)
        self.aplicar.setGeometry(QtCore.QRect(165, 100, 101, 23))
        self.aplicar.clicked.connect(lambda: self.filtros(6))
        self.aplicar.setStyleSheet('QPushButton {\n                border: 2px solid #000000;\n                border-radius: 10px;\n                background-color: #FFFFFF;\n                color: #2E3D48;\n            }\n            QPushButton:pressed {\n                background-color: #2E3D48;\n                color: #FFFFFF;\n            }')
        colo = '\n                            QPushButton {\n\n                            background-color: #FFFFFF;\n                            color: #2E3D48;\n                            border-radius: 10px;\n                            border-color: transparent;\n                            }\n                            QPushButton:hover {\n                            background-color: #c0c0c0;\n                            color: #000000;\n                            }\n                            QPushButton:pressed {\n                                background-color: #2E3D48;\n                                color: #FFFFFF;\n                            }\n                        '
        self.btn_7 = QtWidgets.QPushButton('Últimos 7 dias', self.tela.frame_box)
        self.btn_7.setGeometry(QtCore.QRect(0, 20, 150, 20))
        self.btn_7.setStyleSheet(colo)
        self.btn_personalisa = QtWidgets.QPushButton('Período personalizado >', self.tela.frame_box)
        self.btn_personalisa.setGeometry(QtCore.QRect(0, 40, 150, 20))
        self.btn_personalisa.setStyleSheet(colo)
        self.tela.btn_filtros.clicked.connect(self.abrir_items)
        self.tela.btn_dowload_senso.clicked.connect(self.save_to_excel)
        self.btn_7.clicked.connect(lambda: self.filtros(2))
        self.tela.btnfechar.clicked.connect(lambda: self.filtros(0))
        self.btn_personalisa.clicked.connect(self.abrir_personalisa)
        self.tela.btnfechar.setStyleSheet('QPushButton {    border-top-right-radius: 10px;    border-bottom-right-radius: 10px;    border-top-left-radius: 0px;    border-bottom-left-radius: 0px;    background-color: #FFFFFF;    color: #2E3D48;    border: 2px solid #2E3D48;}QPushButton:pressed {    background-color: #2E3D48;    color: #FFFFFF;}')
        self.tela.btn_filtros.setStyleSheet('QPushButton {\n                border: 2px solid #2E3D48;\n                border-radius: 10px;\n                background-color: #FFFFFF;\n                color: #2E3D48;\n            }\n            QPushButton:pressed {\n                background-color: #2E3D48;\n                color: #FFFFFF;\n            }')
        self.tela.btnfechar.hide()
        self.tela.frame_box.hide()

    def abrir_personalisa(self):
        if self.tela.frame_personalisa.isHidden():
            self.tela.frame_personalisa.show()
            self.tela.frame_box.setStyleSheet('background-color: #FFFFFF; border-top-left-radius: 20px; border-bottom: 1px solid black;')
            self.tela.frame_personalisa.setStyleSheet('\n                        QFrame  {\n                            background-color: #FFFFFF;\n                            border-top-right-radius: 20px;\n                            border-bottom: 1px solid black;\n                            border-left: 1px solid black;\n                        }\n                    ')
        else:  # inserted
            self.tela.frame_personalisa.hide()
            self.tela.frame_box.setStyleSheet('border: 2px solid white; border-radius: 10px; background-color: white;')
            self.tela.frame_personalisa.setStyleSheet('\n                        QFrame  {\n                            background-color: #FFFFFF;\n                            border-top-right-radius: 20px;\n                            border-bottom-right-radius: 20px;\n                            border-left: 1px solid black;\n                        }\n                    ')

    def abrir_items(self):
        if self.tela.frame_box.isHidden():
            self.tela.frame_box.show()
            return
        self.tela.frame_box.hide()
        if not self.tela.frame_personalisa.isHidden():
            self.tela.frame_personalisa.hide()

    def filtros(self, index):
        self.tela.btn_filtros.setStyleSheet('QPushButton {    border-top-right-radius: 0px;    border-bottom-right-radius: 0px;    border-top-left-radius: 10px;    border-bottom-left-radius: 10px;    background-color: #FFFFFF;    color: #2E3D48;    border: 2px solid black;}QPushButton:pressed {    background-color: #2E3D48;    color: #FFFFFF;}')
        from datetime import datetime
        current_month = datetime.now().strftime('%B')
        senso_folder = os.path.join(os.getcwd(), 'senso')
        selected_text = self.combobox_datas.currentText()
        senso_folder = os.path.join(os.getcwd(), 'senso')
        month_year_folder = os.path.join(senso_folder, selected_text)
        if index == 2:
            self.day = 'SEMANA'
            for widget in self.scroll_area.findChildren(QtWidgets.QFrame):
                if widget is not None:
                    widget.deleteLater()
            self.tela.btn_filtros.setText('Ultimos 7 dias')
            self.tela.btn_filtros.setGeometry(QtCore.QRect(80, 0, 120, 23))
            self.tela.btnfechar.show()
            self.load_tables_from_folder(month_year_folder)
        else:  # inserted
            if index == 6:
                self.day = 'PERSONALIZADO'
                self.data_i = self.data_inicio.date()
                self.data_f = self.data_final.date()
                for widget in self.scroll_area.findChildren(QtWidgets.QFrame):
                    if widget is not None:
                        widget.deleteLater()
                self.tela.btn_filtros.setText('Período personalizado')
                self.tela.btn_filtros.setGeometry(QtCore.QRect(80, 0, 120, 23))
                self.tela.btnfechar.show()
                self.load_tables_from_folder(month_year_folder)
            else:  # inserted
                if index == 0:
                    self.day = 'ONTEM'
                    for widget in self.scroll_area.findChildren(QtWidgets.QFrame):
                        if widget is not None:
                            widget.deleteLater()
                    self.tela.btn_filtros.setGeometry(QtCore.QRect(80, 0, 150, 23))
                    self.tela.btn_filtros.setText('▼ Selecione uma Data ')
                    self.tela.btn_filtros.setStyleSheet('QPushButton {\n                border: 2px solid #2E3D48;\n                border-radius: 10px;\n                background-color: #FFFFFF;\n                color: #2E3D48;\n            }\n            QPushButton:pressed {\n                background-color: #2E3D48;\n                color: #FFFFFF;\n            }')
                    self.tela.btnfechar.hide()
                    self.load_tables_from_folder(month_year_folder)

    def listar_pastas_senso(self):
        senso_folder = os.path.join(os.getcwd(), 'senso')
        if not os.path.exists(senso_folder):
            print('A pasta \'senso\' não existe.')
            return
        folders = [folder for folder in os.listdir(senso_folder) if os.path.isdir(os.path.join(senso_folder, folder))]
        return folders

    def atualizar_options_data(self):
        from datetime import datetime
        self.combobox_datas = QComboBox(self.frame)
        current_month_year = datetime.now().strftime('%B_%Y')
        self.combobox_datas.addItem(current_month_year)
        self.combobox_datas.show()
        self.combobox_datas.setGeometry(230, 0, 150, 23)
        self.combobox_datas.setStyleSheet('background-color: rgb(255, 255, 255);border: 2px solid transparent; border-radius: 10px;')
        pastas_senso = self.listar_pastas_senso()
        if pastas_senso:
            for pasta in pastas_senso:
                if pasta!= current_month_year:
                    self.combobox_datas.addItem(f'{pasta}')
        self.combobox_datas.currentIndexChanged.connect(self.on_combobox_changed)

    def on_combobox_changed(self, index):
        selected_text = self.combobox_datas.currentText()
        senso_folder = os.path.join(os.getcwd(), 'senso')
        month_year_folder = os.path.join(senso_folder, selected_text)
        self.load_tables_from_folder(month_year_folder)

    def load_tables_from_folder(self, folder_path):
        if self.frames:
            for frames in self.frames:
                frames.deleteLater()
        self.frames = []
        self.tables = []
        self.tabela_alt.deleteLater()
        self.tabela_alt = QtWidgets.QTableWidget()
        self.tabela_alt.setFixedSize(1383, 680)
        self.cont = 1
        try:
            file_dates = {}
            for file in os.listdir(folder_path):
                if file.endswith('.xlsx'):
                    filepath = os.path.join(folder_path, file)
                    match = re.match(r'\d{2}-\d{2}-\d{4}', file)
                    if match:
                        date_str = match.group()
                        file_dates[filepath] = datetime.datetime.strptime(date_str, '%d-%m-%Y')
            sorted_files = sorted(file_dates.items(), key=lambda x: x[1], reverse=True)
            for file, _ in sorted_files:
                if file.endswith('.xlsx'):
                    filepath = os.path.join(folder_path, file)
                    match = re.match(r'(\d{2})-(\d{2})-(\d{4})', file)
                    if match:
                        self.dia = match.group(1)
                        self.month = match.group(2)
                        self.year = match.group(3)
                    data_atual = QDateTime.currentDateTime()
                    dia = data_atual.date().day() - 7
                    mes = data_atual.date().month()
                    ano = data_atual.date().year()
                    da = int(self.dia)
                    month = int(self.month)
                    year = int(self.year)

                    if self.day == 'PERSONALIZADO':
                        dia = self.data_i.day()
                        mes = self.data_i.month()
                        ano = self.data_i.year()
                        data_inicio = f'{int(dia)}-{int(mes)}-{int(ano)}'
                        dia = self.data_f.day()
                        mes = self.data_f.month()
                        ano = self.data_f.year()
                        data_fim = f'{int(dia)}-{int(mes)}-{int(ano)}'
                        data = f'{self.dia}-{self.month}-{self.year}'
                        if not data_inicio <= data <= data_fim:
                            continue

                    wb = load_workbook(filepath)
                    ws = wb.active
                    max_row = ws.max_row
                    max_column = ws.max_column
                    table_widget = self.add_frame()
                    table_widget.setRowCount(max_row)
                    table_widget.setColumnCount(max_column)
                    conta_linha = table_widget.rowCount() + 2
                    conta_coluna = table_widget.columnCount() + 1 + self.tabela_alt.columnCount()
                    self.tabela_alt.setColumnCount(conta_coluna)
                    self.tabela_alt.setRowCount(conta_linha)
                    data = f'{self.dia}-{self.month}-{self.year}'
                    cont_help = self.cont

                    for row in range(1, max_row + 1):
                        for column in range(1, max_column + 1):
                            cell_value = ws.cell(row=row, column=column).value
                            item = QTableWidgetItem(str(cell_value))
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            table_widget.setItem(row - 1, column - 1, item)
                            item_alt = QTableWidgetItem(str(cell_value))
                            row += 1
                            item_alt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                            self.tabela_alt.setItem(row, self.cont, item_alt)
                            row -= 1
                            self.cont += 1
                        self.cont = cont_help

                    self.cont = cont_help
                    item = QTableWidgetItem(data)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.tabela_alt.setItem(0, int(conta_coluna - table_widget.columnCount() + 1), item)
                    self.tabela_alt.setItem(0, 0, QTableWidgetItem('Data'))

                    columns = ['OCUPAÇÃO', 'BLOQUEIOS', 'CNES', 'OCUPAÇÃO %', 'LEITO/DIA']
                    rows = [
                        '', 'PS|UDC - AGHU', 'PS|CORREDOR - AGHU', 'PS|PEDIATRIA - AGHU', 'PS|UTI - AGHU',
                        '2º LESTE - GRADE', '2º SUL - GRADE', '3º LESTE/CTI - GRADE', '3º NORTE/UCO - GRADE',
                        'MÃE CANGURU - CENSO EBSERH', 'PRÉ-PARTO - CENSO EBSERH', 'MATERNIDADE - CENSO EBSERH',
                        'NEO (UTIN/UCIN) - CENSO EBSERH', '6º LESTE - GRADE', '6º NORTE/CTI PEDIATRICO - GRADE',
                        '7º LESTE - GRADE', '7º NORTE - GRADE', '8º LESTE - GRADE', '8º NORTE - GRADE',
                        '8º SUL - GRADE',
                        '9º LESTE - GRADE', '10º NORTE - GRADE', 'HOSPITAL SÃO GERALDO - AGHU', 'CLÍNICO ADULTO',
                        'CIRÚRGICO ADULTO', 'OBSTÉTRICO', 'PEDIATRICO', 'CTI - ADULTO', 'CTI - PEDIATRICO'
                    ]

                    for row in range(0, len(rows)):
                        item = rows[row]
                        item_alt = QTableWidgetItem(str(item))
                        item_alt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.tabela_alt.setItem(row + 2, 0, item_alt)

                    for column in range(0, len(columns)):
                        item = columns[column]
                        item_alt = QTableWidgetItem(str(item))
                        item_alt.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.tabela_alt.setItem(1, self.cont, item_alt)
                        self.cont += 1
                    self.cont += 1

        except Exception as e:
            print('Erro ao Carregar Tabela', f'Ocorreu um erro ao carregar a tabela:\n{e}')

    def save_to_excel(self):
        default_dir = 'Desktop'
        filename, _ = QFileDialog.getSaveFileName(self.form, 'Salvar Senso', default_dir, 'Arquivos Excel (*.xlsx)')
        if filename:
            from openpyxl.styles import Alignment
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            ws.sheet_format.defaultColWidth = 14
            ws.column_dimensions['A'].width = 35
            for row in range(0, self.tabela_alt.rowCount()):
                for col in range(0, self.tabela_alt.columnCount()):
                    item = self.tabela_alt.item(row, col)
                    if item is None:
                        continue
                    cell_text = item.text()
                    if row == 0 and item.text!= 'Data':
                        cell_text = item.text()
                        next_col = col + 1
                        next_col2 = col + 2
                        next_col3 = col + 3
                        if next_col < self.tabela_alt.columnCount() and next_col2 < self.tabela_alt.columnCount() and (next_col3 < self.tabela_alt.columnCount()):
                            next_item = self.tabela_alt.item(row, next_col)
                            next_item2 = self.tabela_alt.item(row, next_col2)
                            next_item3 = self.tabela_alt.item(row, next_col3)
                            if next_item is not None and next_item2 is not None and (next_item3 is not None) and (next_item.text() == cell_text) and (next_item2.text() == cell_text) and (next_item3.text() == cell_text):
                                ws.merge_cells(start_row=row + 1, start_column=col + 1, end_row=row + 1, end_column=next_col3 + 1)
                    ws.cell(row=row + 1, column=col + 1, value=item.text())
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for cell in ws['1']:
                cell.font = Font(size=16)
            wb.save(filename)

    def add_frame(self):
        frame = QFrame()
        frame_layout = QVBoxLayout(frame)
        frame.setLayout(frame_layout)
        data = QLabel(f'{self.dia}-{self.month}-{self.year}')
        data.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        frame_layout.addWidget(data)
        font = QtGui.QFont()
        font.setPointSize(25)
        data.setFont(font)
        table = QTableWidget()
        frame_layout.addWidget(table)
        columns = ['OCUPAÇÃO', 'BLOQUEIOS', 'CNES', 'OCUPAÇÃO %', 'LEITO/DIA']
        frame.setFixedSize(600, 980)
        rows = ['', 'PS|UDC - AGHU', 'PS|CORREDOR - AGHU', 'PS|PEDIATRIA - AGHU', 'PS|UTI - AGHU', '2º LESTE - GRADE', '2º SUL - GRADE', '3º LESTE/CTI - GRADE', '3º NORTE/UCO - GRADE', 'MÃE CANGURU - CENSO EBSERH', 'PRÉ-PARTO - CENSO EBSERH', 'MATERNIDADE - CENSO EBSERH', 'NEO (UTIN/UCIN) - CENSO EBSERH', '6º LESTE - GRADE', '6º NORTE/CTI PEDIATRICO - GRADE', '7º LESTE - GRADE', '7º NORTE - GRADE', '8º LESTE - GRADE', '8º NORTE - GRADE', '8º SUL - GRADE', '9º LESTE - GRADE', '10º NORTE - GRADE', 'HOSPITAL SÃO GERALDO - AGHU', 'CLÍNICO ADULTO', 'CIRÚRGICO ADULTO', 'OBSTÉTRICO', 'PEDIATRICO', 'CTI - ADULTO', 'CTI - PEDIATRICO']
        table.setRowCount(len(rows))
        if not self.frames:
            table.setVerticalHeaderLabels(rows)
            frame.setFixedSize(830, 980)
        else:  # inserted
            vertical_header = table.verticalHeader()
            vertical_header.setHidden(True)
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)
        table.setColumnWidth(1, 150)
        self.tables.append(table)
        self.frames.append(frame)
        self.scroll_layout.addWidget(frame)
        spacer_item = QSpacerItem(0, 0, QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.scroll_layout.addItem(spacer_item)
        return table