from PyQt6.QtCore import Qt, QSettings, QStandardPaths, QTimer, QTime, QDateTime, QDate
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtGui import QGuiApplication
from PyQt6.QtWidgets import QApplication, QFileDialog, QTableWidget, QTableWidgetItem, QComboBox, QLabel, QVBoxLayout, QWidget, QTimeEdit
from openpyxl import Workbook
from openpyxl import load_workbook
import pymysql
import psycopg2
from database_Demandas import Ui_data_Demanda
from PyQt6.QtCore import QDate
from datetime import datetime
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QDateEdit
from PyQt6.QtCore import QDate
from datetime import datetime, timedelta
FORMATOS_DATA_HORA = [
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S"
    ]

class Ui_Form(object):

    def setupUi(self, Form, frame, tela):
        self.porcentagem = 0
        self.total_hora = 0
        self.data_deman = Ui_data_Demanda()
        self.tela = tela
        self.frame = frame
        self.settings = QSettings('HC', 'SGL')
        self.form = Form
        script_directory = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.AppDataLocation)
        config_file_path = f'{script_directory}/time.ini'
        self.time_load = QSettings(config_file_path, QSettings.Format.IniFormat)
        screen = QGuiApplication.primaryScreen()
        size = screen.size()
        tabela_width = size.width() - 226
        tabela_height = size.height() - 338
        self.tabela_relatorio = QtWidgets.QTableWidget(parent=self.frame)
        self.tabela_relatorio.setGeometry(QtCore.QRect(40, 60, tabela_width, tabela_height))
        self.tabela_relatorio.setObjectName('tableWidget')
        self.tabela_relatorio.setStyleSheet('background-color: rgb(255, 255, 255);gridline-color: black;')
        self.tabela_relatorio.setColumnCount(7)
        self.tabela_relatorio.setRowCount(0)

        self.label_pocentagem_de_meta = QtWidgets.QLabel(parent=self.frame)
        self.label_pocentagem_de_meta.setGeometry(QtCore.QRect(40, 90 + tabela_height, 500, 50))
        self.label_pocentagem_de_meta.show()

        self.label_hora_meta = QtWidgets.QLabel(parent=self.frame)
        self.label_hora_meta.setGeometry(QtCore.QRect(560, 90 + tabela_height, 500, 25))
        self.label_hora_meta.show()

        self.btn_voltar = QtWidgets.QPushButton('VOLTAR',parent=self.frame)
        self.btn_voltar.setGeometry(QtCore.QRect(5, 5, 50, 30))
        self.btn_voltar.show()
        self.btn_voltar.clicked.connect(self.tela.voltar_normal)
        self.btn_voltar.setStyleSheet("QPushButton{\n"
                                      "background-color: rgb(0, 0, 0);\n"
                                      "    color: rgb(255, 255, 255);\n"
                                      "    border-radius:10px;\n"
                                      "}\n"
                                      "QPushButton:hover{\n"
                                      "    background-color: rgb(255, 255, 255);\n"
                                      "    color: rgb(0, 0, 0);\n"
                                      "}")

        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tabela_relatorio.setHorizontalHeaderItem(6, item)

        self.tabela_relatorio.horizontalHeader().setDefaultSectionSize(211)
        self.tabela_relatorio.horizontalHeader().setMinimumSectionSize(65)
        self.tabela_relatorio.verticalHeader().setCascadingSectionResizes(True)
        self.tabela_relatorio.verticalHeader().setDefaultSectionSize(30)
        self.tabela_relatorio.verticalHeader().setMinimumSectionSize(41)

        self.list_tempo_inicial = []
        screen = QGuiApplication.primaryScreen()
        size = screen.size()
        btn_width = size.width() - 166
        btn_width1 = size.height() - 250
        self.btn_dowload_relatotio = QtWidgets.QPushButton('Baixar Relatório', parent=self.frame)
        self.btn_dowload_relatotio.setGeometry(QtCore.QRect(btn_width, btn_width1, 150, 25))
        self.btn_dowload_relatotio.hide()
        self.btn_dowload_relatotio.setToolTip('Baixar Relatório')
        self.btn_dowload_relatotio.clicked.connect(self.save_to_excel)
        self.btn_dowload_relatotio.setStyleSheet('\n                        QPushButton {\n                            border: 2px solid #2E3D48;\n                            border-radius: 10px;\n                            background-color: #FFFFFF;\n                            color: #2E3D48;\n                        }\n\n                        QPushButton:hover {\n                            background-color: #DDDDDD;  /* Change this to your desired hover color */\n                            color: rgb(0, 0, 0);\n                        }\n\n                        QPushButton:pressed {\n                            background-color: #2E3D48;  /* Change this to your desired pressed color */\n                            color: #FFFFFF;\n                        }\n                    ')
        for widget in self.frame.findChildren(QtWidgets.QWidget):
            widget.show()

        self.data_edit = QDateEdit( parent=self.frame)
        self.data_edit.setGeometry(QtCore.QRect(60, 10, 150, 25))
        self.data_edit.setDisplayFormat("MM/yyyy")  # Mostra apenas mês/ano
        self.data_edit.setDate(QDate.currentDate())
        self.data_edit.show()

        self.timer_edit = QtWidgets.QTimeEdit(parent=self.frame)
        self.timer_edit.setGeometry(QtCore.QRect(240, 10, 150, 25))
        self.timer_edit.setDisplayFormat("HH:mm:ss")
        self.timer_edit.setTime(QtCore.QTime(4, 59, 59))
        self.timer_edit.show()

        self.numero_edit = QtWidgets.QDoubleSpinBox(parent=self.frame)
        self.numero_edit.setGeometry(QtCore.QRect(420, 10, 150, 25))
        self.numero_edit.setSuffix(" %")  # Sufixo de porcentagem
        self.numero_edit.setValue(75.0)  # Valor inicial
        self.numero_edit.setRange(0.0, 100.0)  # Faixa de 0% a 100%
        self.numero_edit.show()

        self.data_edit.dateChanged.connect(self.atualizar_relatorio)
        self.timer_edit.timeChanged.connect(self.atualizar_relatorio)
        self.numero_edit.valueChanged.connect(self.atualizar_relatorio)

        self.retranslateUi(Form)
        self.atualizar_relatorio()
        self.timer = QTimer()
        #self.timer.timeout.connect(self.atualizar_tempo)
        self.timer.start(1000)
        self.tempo_decorrido = 0
        #self.tabela_relatorio.horizontalHeader().sectionClicked.connect(self.mostrar_showTimeEdit)
        self.timeEdit = None
        self.timeEdit_10 = None
        #self.atualizar_ocupacao()

    def mostrar_showTimeEdit(self, logicalIndex):
        if logicalIndex == 9:
            if self.timeEdit is not None:
                self.tabela_relatorio.removeCellWidget(0, logicalIndex)
                self.time_load.setValue('time', self.timeEdit.time().toString('hh:mm'))
                self.analise_alta_meta()
                self.timeEdit = None
            else:
                self.timeEdit = QTimeEdit()
                self.timeEdit.setMaximumTime(QTime(99, 59, 59))
                self.timeEdit.setDisplayFormat('hh:mm')
                time = self.carregar_tempo()
                self.timeEdit.setTime(time)
                self.tabela_relatorio.setCellWidget(0, logicalIndex, self.timeEdit)
        if logicalIndex == 10:
            if self.timeEdit_10 is not None:
                self.tabela_relatorio.removeCellWidget(0, logicalIndex)
                self.time_load.setValue('time_2', self.timeEdit_10.time().toString('hh:mm'))
                self.calcular_soma()
                self.timeEdit_10 = None
            else:
                self.timeEdit_10 = QTimeEdit()
                self.timeEdit_10.setMaximumTime(QTime(99, 59, 59))
                self.timeEdit_10.setDisplayFormat('hh:mm')
                time = self.carregar_tempo_2()
                self.timeEdit_10.setTime(time)
                self.tabela_relatorio.setCellWidget(0, logicalIndex, self.timeEdit_10)

    def carregar_tempo(self):
        if self.time_load.contains('time'):
            time_str = self.time_load.value('time')
            time_parts = time_str.split(':')
            hours, minutes = map(int, time_parts)
            return QTime(hours, minutes)
        return QTime(0, 0)

    def carregar_tempo_2(self):
        if self.time_load.contains('time_2'):
            time_str = self.time_load.value('time_2')
            time_parts = time_str.split(':')
            hours, minutes = map(int, time_parts)
            return QTime(hours, minutes)
        return QTime(0, 0)

    def atualizar_tempo(self):
        for row, tempo in self.list_tempo_inicial:
            tempo_atual = QDateTime.currentDateTime()
            diferenca = tempo.secsTo(tempo_atual)
            tempo_decorrido = QTime(0, 0).addSecs(diferenca).toString('hh:mm:ss')
            item = QtWidgets.QTableWidgetItem(tempo_decorrido)
            print(item.text())
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.tabela_relatorio.setItem(row, 5, item)

    def descobrir_nome_coluna(self, nome, coluna):
        for col in range(self.tabela_relatorio.columnCount()):
            item_pac = self.tabela_relatorio.horizontalHeaderItem(col)
            if item_pac is not None:
                if item_pac.text() == nome:
                    return col
        return coluna
    def atualizar_relatorio(self):
        contador_Sim = 0
        contador_tempo = 0
        coluna_data_alta = self.data_deman.pegar_coluna_Demanda(self, 'alta_cti', 'DATA DA ALTA')
        coluna_hora_alta = self.data_deman.pegar_coluna_Demanda(self, 'alta_cti', 'HORA DE SOLICITAÇÃO DA ALTA')
        coluna_nome_alta = self.data_deman.pegar_coluna_Demanda(self, 'alta_cti', 'NOME DO PACIENTE')
        coluna_conf_alta = self.data_deman.pegar_coluna_Demanda(self, 'alta_cti', 'DATA E HORÁRIO DA CONFIRMAÇÃO DA ALTA')
        coluna_stat_alta = self.data_deman.pegar_coluna_Demanda(self, 'alta_cti', 'STATUS DA SOLICITAÇÃO')

        coluna_data = "col" if coluna_data_alta == 0 else f"col{coluna_data_alta}"
        coluna_hora = "col" if coluna_hora_alta == 0 else f"col{coluna_hora_alta}"
        coluna_nome = "col" if coluna_nome_alta == 0 else f"col{coluna_nome_alta}"
        coluna_conf = "col" if coluna_conf_alta == 0 else f"col{coluna_conf_alta}"
        coluna_stat = "col" if coluna_stat_alta == 0 else f"col{coluna_stat_alta}"

        dado_data = self.data_deman.pegar_Dados_Demanda(self,'alta_cti', coluna_data)
        dado_hora = self.data_deman.pegar_Dados_Demanda(self,'alta_cti', coluna_hora)
        dado_nome = self.data_deman.pegar_Dados_Demanda(self,'alta_cti', coluna_nome)
        dado_conf = self.data_deman.pegar_Dados_Demanda(self,'alta_cti', coluna_conf)
        dado_stat = self.data_deman.pegar_Dados_Demanda(self,'alta_cti', coluna_stat)

        print('mano',dado_conf,dado_data,dado_hora,dado_nome,dado_stat)

        x = self.timer_edit.time()
        coluna_nome_tabela = self.descobrir_nome_coluna('NOME DO PACIENTE',None)
        coluna_data_tabela = self.descobrir_nome_coluna('DATA DA ALTA',None)
        coluna_hora_tabela = self.descobrir_nome_coluna('HORA DE SOLICITAÇÃO DA ALTA',None)
        coluna_conf_tabela = self.descobrir_nome_coluna('DATA E HORÁRIO DA CONFIRMAÇÃO DA ALTA',None)
        coluna_stat_tabela = self.descobrir_nome_coluna('STATUS DA SOLICITAÇÃO',None)
        coluna_temo_libera = self.descobrir_nome_coluna('TEMPO DE LIBERAÇÃO DE LEITO PARA ALTAS DO CTI:         ', None)

        hora_formatada = self.timer_edit.time()
        x = hora_formatada.toString("HH:mm:ss")
        coluna_analise_libera = self.descobrir_nome_coluna(f"LIBERADO ANTES DE {x} ", None)

        self.tabela_relatorio.setRowCount(0)
        for cont , nome in enumerate(dado_nome):
            if len(dado_nome) < 1 or dado_nome[cont] is None:
                continue

            if len(dado_conf) < 1 or dado_conf[cont] is None:
                data_str = ''
            else:
                data_str = dado_conf[cont]

            if len(dado_data) < 1 or dado_data[cont] is None:
                dado_data_str = ''
            else:
                dado_data_str = dado_data[cont]

            if len(dado_hora) < 1 or dado_hora[cont] is None:
                dado_hora_str = ''
            else:
                dado_hora_str = dado_hora[cont]

            if len(dado_stat) < 1 or dado_stat[cont] is None:
                dado_stat_str = ''
            else:
                dado_stat_str = dado_stat[cont]

            data_usuario = self.data_edit.date()
            mes_usuario = data_usuario.month()
            ano_usuario = data_usuario.year()

            if data_str != '':
                if isinstance(data_str, str):
                    try:
                        data_dt = datetime.strptime(data_str, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            data_dt = datetime.strptime(data_str, "%d/%m/%Y %H:%M")
                        except ValueError:
                            print(f"Formato de data inválido: {data_str}")
                            continue  # ou levante erro, dependendo do seu caso
                else:
                    data_dt = data_str

                mes_lista = data_dt.month
                ano_lista = data_dt.year

                if mes_usuario != mes_lista or ano_usuario != ano_lista:
                    continue

            row = self.tabela_relatorio.rowCount()
            self.tabela_relatorio.insertRow(row)

            def fazer_item(text):
                item = QtWidgets.QTableWidgetItem(text)
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                if item is None:
                    item = QtWidgets.QTableWidgetItem(str(''))
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                return item

            self.tabela_relatorio.setItem(row, coluna_nome_tabela, fazer_item(nome))
            self.tabela_relatorio.setItem(row, coluna_data_tabela, fazer_item(dado_data_str))
            self.tabela_relatorio.setItem(row, coluna_hora_tabela, fazer_item(dado_hora_str))
            self.tabela_relatorio.setItem(row, coluna_conf_tabela, fazer_item(data_str))
            self.tabela_relatorio.setItem(row, coluna_stat_tabela, fazer_item(dado_stat_str))

            diferenca = self.calcular_diferenca(dado_data_str,dado_hora_str,data_str)
            if diferenca is not None:
                print(f"Diferença: {diferenca}")
                self.tabela_relatorio.setItem(row, coluna_temo_libera, fazer_item(diferenca))

                def string_para_segundos(hhmmss: str) -> int:
                    h, m, s = map(int, hhmmss.split(":"))
                    return h * 3600 + m * 60 + s

                contador_tempo += string_para_segundos(diferenca)

                def string_para_timedelta(horario_str):
                    h, m, s = map(int, horario_str.split(":"))
                    return timedelta(hours=h, minutes=m, seconds=s)

                diferenca_tempo = string_para_timedelta(diferenca)
                tempo_qt = self.timer_edit.time()
                tempo_limite = timedelta(hours=tempo_qt.hour(), minutes=tempo_qt.minute(), seconds=tempo_qt.second())

                if diferenca_tempo > tempo_limite:
                    print("Passou do tempo limite.")
                    self.tabela_relatorio.setItem(row, coluna_analise_libera, fazer_item('NÃO'))
                else:
                    print("Está dentro do limite.")
                    self.tabela_relatorio.setItem(row, coluna_analise_libera, fazer_item('SIM'))
                    contador_Sim+=1
            else:
                print("Não foi possível calcular a diferença.")

        valor = self.numero_edit.value()
        if self.tabela_relatorio.rowCount() > 0:
            if contador_tempo > 0:
                contador_tempo = contador_tempo / self.tabela_relatorio.rowCount()
                horas = int(contador_tempo // 3600)
                minutos = int((contador_tempo % 3600) // 60)
                segundos = int(contador_tempo % 60)
                tempo_formatado = f"{horas:02}:{minutos:02}:{segundos:02}"

                self.total_hora = tempo_formatado
                self.label_hora_meta.setText(f'MÉDIA DE LIBERAÇÃO DE LEITO PARA AS ALTAS DO CTI POR MÊS: {tempo_formatado}')
            self.label_pocentagem_de_meta.setText(f'% Liberação de Alta \n- Meta: {valor} % - \n- Porcentagem Real: {(contador_Sim / self.tabela_relatorio.rowCount())*100} % -')
            self.porcentagem = (contador_Sim / self.tabela_relatorio.rowCount())*100

    def parse_datetime_flexivel(self,data_str):
        for formato in FORMATOS_DATA_HORA:
            try:
                return datetime.strptime(data_str.strip(), formato)
            except (ValueError, AttributeError):
                continue
        print(f"Formato de data inválido: {data_str}")
        return None

    def calcular_diferenca(self,data_alta,hora_alta,data_confirmacao):
        # Pegando os dados da linha

        # Concatenando data + hora para formar o datetime da alta
        datetime_alta = self.parse_datetime_flexivel(f"{data_alta} {hora_alta}")
        datetime_confirmacao = self.parse_datetime_flexivel(data_confirmacao)

        if datetime_alta and datetime_confirmacao:
            diferenca = datetime_confirmacao - datetime_alta
            total_segundos = int(diferenca.total_seconds())

            horas = total_segundos // 3600
            minutos = (total_segundos % 3600) // 60
            segundos = total_segundos % 60

            return f"{horas:02}:{minutos:02}:{segundos:02}"
        else:
            return None

    def analise_alta_meta(self):
        for row in range(self.tabela_relatorio.rowCount()):
            time2 = QTimeEdit()
            time2.setDisplayFormat('hh:mm:ss')
            time1 = self.tabela_relatorio.item(row, 5).text()
            time = self.carregar_tempo()
            time2.setTime(time)
            if time1 <= time2.time().toString('hh:mm:ss'):
                item = QtWidgets.QTableWidgetItem('Abaixo da meta')
            else:
                item = QtWidgets.QTableWidgetItem('1')
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.tabela_relatorio.setItem(row, 9, item)
            self.calcular_soma()

    def calcular_soma(self):
        soma = 0
        altas = 0
        altas_day = 0
        average_time = QTime(0, 0, 0)
        current_time = QDateTime.currentDateTime()
        for row in range(self.tabela_relatorio.rowCount()):
            item = self.tabela_relatorio.item(row, 9)
            if item is not None and item.text() == '1':
                soma += int(self.tabela_relatorio.item(row, 9).text())
            if self.tabela_relatorio.item(row, 6) is not None and ('RESERVA' in self.tabela_relatorio.item(row, 6).text() or 'OCUPADO' in self.tabela_relatorio.item(row, 6).text()):
                selected_date_text = self.tabela_relatorio.item(row, 0).text()
                selected_date = QDateTime.fromString(selected_date_text, 'dd/MM/yyyy')
                current_date = QDateTime.currentDateTime()
                if selected_date.date().month() == current_date.date().month():
                    total_seconds = 0
                    row_count = self.tabela_relatorio.rowCount()
                    for row in range(row_count):
                        item = self.tabela_relatorio.item(row, 5)
                        time_str = item.text()
                        time = QTime.fromString(time_str, 'hh:mm:ss')
                        total_seconds += time.hour() * 3600 + time.minute() * 60 + time.second()
                    average_seconds = int(total_seconds / row_count)
                    average_time = QTime(0, 0, 0).addSecs(average_seconds)
        for row in range(self.tabela_relatorio.rowCount()):
            selected_date_text = self.tabela_relatorio.item(row, 0).text()
            selected_time_text = self.tabela_relatorio.item(row, 1).text()
            selected_date_time = QDateTime.fromString(f'{selected_date_text} {selected_time_text}', 'dd/MM/yyyy hh:mm:ss')
            subtracted_time = self.subtract_time()
            if current_time >= selected_date_time and selected_date_time >= subtracted_time:
                altas += 1
                item = QtWidgets.QTableWidgetItem('1')
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.tabela_relatorio.setItem(row, 10, item)
            else:
                item = QtWidgets.QTableWidgetItem('0')
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.tabela_relatorio.setItem(row, 10, item)
            selected_date_text = self.tabela_relatorio.item(row, 0).text()
            selected_date = QDateTime.fromString(selected_date_text, 'dd/MM/yyyy')
            current_date = QDateTime.currentDateTime()
            if selected_date.date().day() == current_date.date().day():
                altas_day += 1
                continue
            item = QtWidgets.QTableWidgetItem('ALTA ANTIGA')
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.tabela_relatorio.setItem(row, 10, item)
        row = self.tabela_relatorio.rowCount()
        valor = '{:.2f} %'.format(soma / row * 100)
        item = QtWidgets.QTableWidgetItem(f'%  DE ALTAS ABAIXO DA META: {valor}')
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabela_relatorio.setHorizontalHeaderItem(9, item)
        valor = average_time.toString('hh:mm:ss')
        item = QtWidgets.QTableWidgetItem(f'TEMPO MÉDIO NO MÊS PARA LIBERAÇÃO DE LEITOS: {valor}')
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabela_relatorio.setHorizontalHeaderItem(5, item)
        self.timeEdit_10 = QTimeEdit()
        self.timeEdit_10.setMaximumTime(QTime(99, 59, 59))
        time = self.carregar_tempo_2()
        self.timeEdit_10.setTime(time)
        valor = self.timeEdit_10.time().toString('hh:mm:ss')
        if altas_day != 0:
            valor_2 = '{:.2f} %'.format(altas / altas_day * 100)
        else:
            valor_2 = '0.00 %'
        item = QtWidgets.QTableWidgetItem(f'%  DE ALTAS EM {valor} HORAS: {valor_2}')
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabela_relatorio.setHorizontalHeaderItem(10, item)

    def atualizar_ocupacao(self):
        self.cont_ocupado_cti = 0
        self.total_cti = 0
        self.cont_ocupado_enf = 0
        self.total_enf = 0
        list = ['CTI PEDIÁTRICO - 06N', 'UNIDADE DE INTERNAÇÃO CORONARIANA - 03N', 'UTI - PRONTO SOCORRO', 'CTI ADULTO - 03L', 'UNIDADE DE INTERNAÇÃO - 06L', 'UNIDADE DE INTERNAÇÃO - 10N', 'UNIDADE DE INTERNAÇÃO - 02L', 'UNIDADE DE INTERNAÇÃO - 02S', 'UNIDADE DE INTERNAÇÃO - 07L', 'UNIDADE DE INTERNAÇÃO - 07N', 'UNIDADE DE INTERNAÇÃO - 08S', 'UNIDADE DE INTERNAÇÃO - 08L', 'UNIDADE DE INTERNAÇÃO - 08N', 'UNIDADE DE INTERNAÇÃO - 09L']
        conexao = pymysql.connect(host='10.36.0.32', user='sglHC2024', password='S4g1L81', database='sgl')
        cursor = conexao.cursor()
        comando = 'SELECT idNew_GRADES FROM New_GRADES '
        cursor.execute(comando)
        rows = cursor.fetchall()
        for row in rows:
            list.append(row)
        cursor.close()
        conexao.close()
        for nome in list:
            self.conferir_ocupacao(nome)
        if self.total_cti == 0:
            valor = '0.00 %'
        else:
            valor = '{:.2f} %'.format(self.cont_ocupado_cti / self.total_cti * 100)
        item = QtWidgets.QTableWidgetItem(f"TAXA DE OCUPAÇÃO DOS CTI'S: {valor}")
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabela_relatorio.setHorizontalHeaderItem(7, item)
        if self.total_enf == 0:
            valor = '0.00 %'
        else:
            valor = '{:.2f} %'.format(self.cont_ocupado_enf / self.total_enf * 100)
        item = QtWidgets.QTableWidgetItem(f'TAXA DE OCUPAÇÃO DAS ENFERMARIAS: {valor}')
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.tabela_relatorio.setHorizontalHeaderItem(8, item)
        self.tela.timer.stop()
        self.tela.timer_post.stop()
        self.tela.timer_mysql.stop()

    def conferir_ocupacao(self, nome):
        import psycopg2
        connection = psycopg2.connect(user='ugen_integra', password='aghuintegracao', host='10.36.2.35', port='6544', database='dbaghu')
        cursor = connection.cursor()
        cursor.execute(f"SELECT leitos.lto_id, leitos.leito, uni.ind_unid_cti, uni.ind_unid_internacao FROM agh.ain_leitos leitos INNER JOIN agh.AGH_UNIDADES_FUNCIONAIS uni ON uni.seq = leitos.unf_seq WHERE descricao = '{nome}'")
        rows = cursor.fetchall()
        self.lista_sexo = []
        for row in rows:
            sem_hifen = row[0].split('-')[0]
            semzero = row[1].lstrip('0')
            dados = f'{sem_hifen}_{semzero}'
            if row[3] == 'S' or row[2] == 'S':
                self.conferir_ocupacao_cti(dados)
            else:
                self.conferir_ocupacao_enfermaria(dados)
        cursor.close()
        connection.close()

    def conferir_ocupacao_cti(self, dados):
        conexao = pymysql.connect(host='10.36.0.32', user='sglHC2024', password='S4g1L81', database='sgl')
        cursor = conexao.cursor()
        comando = f'SELECT STATUS_DO_LEITO FROM GRADE WHERE idGRADE = "{dados}"'
        cursor.execute(comando)
        rows = cursor.fetchall()
        for row in rows:
            if row[0] == 'OCUPADO':
                self.cont_ocupado_cti += 1
            self.total_cti += 1
        cursor.close()
        conexao.close()

    def conferir_ocupacao_enfermaria(self, dados):
        conexao = pymysql.connect(host='10.36.0.32', user='sglHC2024', password='S4g1L81', database='sgl')
        cursor = conexao.cursor()
        comando = f"SELECT SEXO DA ENFERMARIA FROM GRADE WHERE idGRADE = '{dados}'"
        cursor.execute(comando)
        rows = cursor.fetchall()
        for row in rows:
            if row[0] == 'OCUPADO':
                self.cont_ocupado_enf += 1
            self.total_enf += 1
        cursor.close()
        conexao.close()

    def subtract_time(self):
        current_date_time = QDateTime.currentDateTime()
        current_time = current_date_time.time()
        current_date = current_date_time.date()
        self.timeEdit = QTimeEdit()
        self.timeEdit.setMaximumTime(QTime(99, 59, 59))
        time = self.carregar_tempo_2()
        self.timeEdit.setTime(time)
        subtracted_time = self.timeEdit.time()
        new_time = QTime((current_time.hour() - subtracted_time.hour()) % 24, (current_time.minute() - subtracted_time.minute()) % 60, (current_time.second() - subtracted_time.second()) % 60)
        if current_time < subtracted_time:
            current_date = current_date.addDays(-1)
        result_date_time = QDateTime(current_date, new_time)
        return result_date_time

    def save_to_excel(self):
        default_dir = 'C:/Users/User/Desktop'
        filename, _ = QFileDialog.getSaveFileName(self.form, 'Salvar Relatório', default_dir, 'Arquivos Excel (*.xlsx)')
        rows = [row for row, _ in self.list_tempo_inicial]
        if filename:
            from openpyxl.styles import Alignment
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            valor = self.numero_edit.value()

            widths = [35, 45, 35, 45, 60, 80, 80, 80, 90, 60, 60]
            titulo1= 'MÉDIA DE LIBERAÇÃO DE LEITO PARA AS ALTAS DO CTI POR MÊS'
            titulo2 =  f'% Liberação de Alta - Meta {valor}%'

            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            for col in range(0, self.tabela_relatorio.columnCount()):
                item = self.tabela_relatorio.horizontalHeaderItem(col)
                ws.cell(row=1, column=col + 1, value=item.text())
            ws.cell(row=1, column=col + 2, value=titulo1)
            ws.cell(row=1, column=col + 3, value=titulo1)


            for row in range(0, self.tabela_relatorio.rowCount()):
                for col in range(0, self.tabela_relatorio.columnCount()):
                    item = self.tabela_relatorio.item(row, col)
                    if item is not None:
                        item = item.text()
                        if col == 5 and row in rows:
                            item = str('LEITO NÃO LIBERADO')
                        ws.cell(row=row + 2, column=col + 1, value=item)

            ws.cell(row= 2, column=col + 2, value=self.total_hora)
            ws.cell(row= 2, column=col + 3, value=self.porcentagem)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for cell in ws['1']:
                cell.font = Font(size=16)
            wb.save(filename)

    def retranslateUi(self, Form):
        hora_formatada = self.timer_edit.time()
        x = hora_formatada.toString("HH:mm:ss")

        _translate = QtCore.QCoreApplication.translate
        item = self.tabela_relatorio.horizontalHeaderItem(0)
        item.setText(_translate('Form', 'DATA DA ALTA'))
        item = self.tabela_relatorio.horizontalHeaderItem(1)
        item.setText(_translate('Form', 'HORA DE SOLICITAÇÃO DA ALTA'))
        item = self.tabela_relatorio.horizontalHeaderItem(2)
        item.setText(_translate('Form', 'NOME DO PACIENTE'))
        item = self.tabela_relatorio.horizontalHeaderItem(3)
        item.setText(_translate('Form', 'STATUS DA SOLICITAÇÃO'))
        item = self.tabela_relatorio.horizontalHeaderItem(4)
        item.setText(_translate('Form', 'DATA E HORÁRIO DA CONFIRMAÇÃO DA ALTA'))
        item = self.tabela_relatorio.horizontalHeaderItem(5)
        item.setText(_translate('Form', 'TEMPO DE LIBERAÇÃO DE LEITO PARA ALTAS DO CTI:         '))
        item = self.tabela_relatorio.horizontalHeaderItem(6)
        item.setText(_translate('Form', f"LIBERADO ANTES DE {x} "))

        for coluna in range(1, self.tabela_relatorio.columnCount()):
            item_pac = self.tabela_relatorio.horizontalHeaderItem(coluna).text()
            font_tabela = self.tabela_relatorio.font()
            text_width = font_tabela.pointSize() * len(item_pac)
            self.tabela_relatorio.setColumnWidth(coluna, text_width + 100)